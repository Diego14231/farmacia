"""
Clasificador de inventario de farmacia contra el Registro Sanitario del ISP (Chile).

Qué hace:
  Toma TU MISMO inventario (con todas sus columnas originales: Código,
  Producto, P. Costo, P. Venta, etc.) y le agrega 4 columnas nuevas al final:

    - Es Medicamento        -> SI / NO
    - Requiere Receta       -> "NO (venta directa)" / "SÍ - Receta Simple" /
                               "SÍ - Receta Retenida" / "SÍ - Receta Cheque" /
                               "SÍ - Receta Retenida con Control de Existencia" /
                               "NO DETERMINADO" (si es medicamento pero no se
                               pudo confirmar el tipo) / vacío (si no es
                               medicamento)
    - Revisar Manual        -> SI / NO (para que la Química Farmacéutica
                               filtre y revise solo estas filas primero)
    - Detalle               -> registro ISP + principio activo + laboratorio
                               si se encontró, o candidatos alternativos si
                               quedó ambiguo, o el motivo si hubo un error

  El resultado es UN SOLO ARCHIVO que se ve igual a tu Excel de siempre, con
  esas 4 columnas nuevas al final -- no un archivo aparte con otra estructura.

  Cada producto se busca en https://registrosanitario.ispch.gob.cl/ (portal
  público del ISP). Todo queda como "sugerido": la Química Farmacéutica del
  establecimiento revisa y confirma, no reemplaza su firma (requisito legal
  del ISP -- ver PLAN-FARMACIA-ONLINE.md sección 9).

Requisitos:
  pip install playwright openpyxl
  playwright install chromium

Uso:
  python clasificar_medicamentos.py inventario1.xlsx salida_clasificacion.csv

  Se puede interrumpir (Ctrl+C) y volver a correr con el mismo archivo de
  salida: continúa donde quedó (no repite productos ya procesados).

  Al final, convertir a Excel con:
  python csv_a_excel_seguro.py salida_clasificacion.csv salida_clasificacion.xlsx

Nota sobre velocidad: es scraping de un portal de gobierno, así que corre a
un ritmo deliberadamente conservador (pausa entre requests). Para las ~4.986
filas del inventario real, calcular varias horas de ejecución. No se hizo
para ser rápido, se hizo para no abusar del sitio del ISP.
"""

import csv
import os
import re
import sys
import time

from openpyxl import load_workbook
from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout

BASE_URL = "https://registrosanitario.ispch.gob.cl/"
REQUEST_DELAY_SECONDS = 1.5
MAX_RETRIES = 3
# Si un término (ej. "ACEITE") devuelve más resultados que esto, no se leen
# uno por uno -- de todas formas terminarían en revisión manual por
# ambiguos, y leer cientos de filas es lo que dejaba el proceso pegado
# varios minutos en un solo producto.
MAX_CANDIDATOS_A_LEER = 40

COLUMNAS_NUEVAS = ["Es Medicamento", "Requiere Receta", "Revisar Manual", "Detalle"]

# Los 5 valores oficiales, en orden de frecuencia esperada (para minimizar
# consultas promedio -- ver PLAN-FARMACIA-ONLINE.md sección 8.5).
CONDICIONES_VENTA = [
    "Directa",
    "Receta Simple",
    "Receta Retenida",
    "Receta Cheque",
    "Receta Retenida con Control de Existencia",
]

ETIQUETA_REQUIERE_RECETA = {
    "Directa": "NO (venta directa)",
    "Receta Simple": "SÍ - Receta Simple",
    "Receta Retenida": "SÍ - Receta Retenida",
    "Receta Cheque": "SÍ - Receta Cheque",
    "Receta Retenida con Control de Existencia": "SÍ - Receta Retenida con Control de Existencia",
}

# Señal heurística "esto suena a medicamento" -- solo para decidir si un
# producto SIN match amerita revisión manual igual (no para clasificar nada).
# Se separa en dos partes a propósito:
#   - FORM_HINT_RE: palabras de forma farmacéutica (con \b, seguro porque no
#     suelen ir pegadas a un número).
#   - DOSIS_RE: la dosis en sí. OJO -- en los nombres reales del inventario la
#     dosis casi siempre viene PEGADA al número ("27.5MCG", "1000MG"), así que
#     no se puede usar \b antes de la unidad (\b no marca límite entre un
#     dígito y una letra, ambos son \w) o se pierden casos reales como
#     "FREMAVAL FLUTICASONA Cenabast 27.5MCG/DOSIS".
# También se excluyen a propósito "ML" y "G" sueltos como palabra completa --
# aparecen en cualquier perfume/shampoo por el volumen del envase y generan
# demasiados falsos positivos.
FORM_HINT_RE = re.compile(
    r"\b(COMP(?:RIMIDOS?)?|TAB(?:LETAS?)?|CAPS(?:ULAS?)?|JARABE|GOTAS|"
    r"INYECTABLE|OVULO|SUPOSITORIO|AMPOLLA|GRAGEA|PERLAS?|SOBRES?|PARCHE)\b",
    re.IGNORECASE,
)

DOSIS_RE = re.compile(r"(\d+[.,]?\d*)\s?(MG|MCG|UI)\b", re.IGNORECASE)
FORMA_KEYWORDS = [
    "comprimido", "jarabe", "gota", "capsula", "cápsula", "crema", "unguento",
    "ovulo", "supositorio", "ampolla", "solucion", "solución", "gragea",
    "perla", "inyectable", "spray", "polvo", "suspension", "suspensión",
]

# Tokens que marcan "acá termina el nombre y empieza la dosis/presentación" --
# se usan para cortar el nombre sucio del inventario a un término de búsqueda
# limpio (ver limpiar_para_busqueda). Ej: "GLAUPAX XR 1000MG X 30 COMP" corta
# en "1000MG", no antes, porque "XR" es parte del nombre comercial.
TOKEN_DE_CORTE_RE = re.compile(
    r"^\d|^(X|MG|MCG|UI|COMP(?:RIMIDOS?)?|TAB(?:LETAS?)?|CAPS(?:ULAS?)?|"
    r"SOBRES?|UN|U)\.?$",
    re.IGNORECASE,
)


def suena_a_medicamento(nombre):
    return bool(DOSIS_RE.search(nombre) or FORM_HINT_RE.search(nombre))


def extraer_dosis(nombre):
    m = DOSIS_RE.search(nombre)
    if not m:
        return None
    valor, unidad = m.groups()
    return f"{valor}{unidad}".lower().replace(",", ".")


def extraer_formas(nombre):
    nombre_low = nombre.lower()
    return {f for f in FORMA_KEYWORDS if f in nombre_low}


def limpiar_para_busqueda(nombre, max_palabras=1):
    """Extrae un término de búsqueda corto a partir del nombre sucio del
    inventario -- ej. "ABRILAR 35 MG/5ML JARABE 100 ML (HEDERA HELIX)" -> "ABRILAR".

    Se probó en el piloto que el sitio del ISP no matchea si se le manda el
    nombre completo (con dosis, presentación, marca del laboratorio, etc.)
    -- hay que mandarle solo el nombre/marca y filtrar los resultados
    despues por dosis/forma en Python (ver clasificar_producto).
    """
    sin_parentesis = re.sub(r"\([^)]*\)", " ", nombre)
    tokens = sin_parentesis.strip().split()
    if not tokens:
        return nombre.strip()
    salida = []
    for t in tokens:
        if TOKEN_DE_CORTE_RE.match(t) and salida:
            break
        salida.append(t)
        if len(salida) >= max_palabras:
            break
    return " ".join(salida) if salida else tokens[0]


def leer_inventario(path_excel):
    """Devuelve (header_original, filas) preservando TODAS las columnas y el
    orden original del Excel de la farmacia -- este script solo agrega
    columnas nuevas al final, no reemplaza ni reordena nada de lo que ya
    existe."""
    wb = load_workbook(path_excel, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    header = [str(h).strip() for h in rows[0]]
    idx_codigo = header.index("Código") if "Código" in header else 0
    idx_producto = header.index("Producto") if "Producto" in header else 1

    filas = []
    for r in rows[1:]:
        if r[idx_producto] is None:
            continue
        fila = {header[i]: (r[i] if r[i] is not None else "") for i in range(len(header))}
        fila["_codigo"] = str(r[idx_codigo])
        fila["_producto"] = str(r[idx_producto]).strip()
        filas.append(fila)
    return header, filas


def cargar_ya_procesados(path_csv, columna_codigo):
    if not os.path.exists(path_csv):
        return set()
    try:
        with open(path_csv, newline="", encoding="utf-8-sig") as f:
            return {row[columna_codigo] for row in csv.DictReader(f)}
    except PermissionError:
        _error_archivo_abierto(path_csv)


def abrir_csv_para_append(path_csv, fieldnames):
    nuevo = not os.path.exists(path_csv)
    try:
        f = open(path_csv, "a", newline="", encoding="utf-8-sig")
    except PermissionError:
        _error_archivo_abierto(path_csv)
    writer = csv.DictWriter(f, fieldnames=fieldnames)
    if nuevo:
        writer.writeheader()
    return f, writer


def _error_archivo_abierto(path_csv):
    print(
        f"\nNo se puede abrir '{path_csv}' -- lo más probable es que esté "
        "abierto en Excel ahora mismo.\n"
        "Ciérralo en Excel (o cierra Excel completo) y vuelve a correr el "
        "mismo comando -- no perdiste nada, retoma donde quedó.\n"
    )
    sys.exit(1)


def parece_navegador_muerto(error):
    """True si el error significa que el navegador/pestaña se cerró o
    crasheó (no un error normal de búsqueda) -- hay que relanzar el
    navegador entero, no solo reintentar la misma llamada."""
    msg = str(error).lower()
    return any(
        s in msg
        for s in (
            "closed",
            "crashed",
            "target page",
            "target closed",
            "connection closed",
            "browser has been closed",
        )
    )


class RegistroSanitarioClient:
    """Envoltorio del portal registrosanitario.ispch.gob.cl vía navegador real.

    Se probó (ver pilot) que reconstruir a mano el postback ASP.NET con
    requests planas falla de forma consistente ("Validation of viewstate MAC
    failed"); un navegador real sí funciona de forma confiable, así que este
    cliente usa Playwright en vez de requests/urllib.
    """

    def __init__(self, page):
        self.page = page

    def _ir_a_busqueda_por_nombre(self):
        self.page.goto(BASE_URL, timeout=30000)
        self.page.check("#ctl00_ContentPlaceHolder1_chkTipoBusqueda_0")
        self.page.wait_for_selector(
            "#ctl00_ContentPlaceHolder1_txtNombreProducto", timeout=10000
        )

    def _ir_a_busqueda_por_registro_y_condicion(self):
        self.page.goto(BASE_URL, timeout=30000)
        self.page.check("#ctl00_ContentPlaceHolder1_chkTipoBusqueda_3")  # Número de Registro
        self.page.check("#ctl00_ContentPlaceHolder1_chkTipoBusqueda_5")  # Condición de Venta
        self.page.wait_for_selector(
            "#ctl00_ContentPlaceHolder1_txtNumeroRegistro", timeout=10000
        )

    def _leer_resultados(self):
        cuerpo = self.page.inner_text("body")
        m = re.search(r"Registros Encontrados\s*:\s*(\d+)", cuerpo)
        total = int(m.group(1)) if m else 0

        filas = []
        # Si un término genérico ("ACEITE", "CREMA", etc.) devuelve cientos de
        # resultados, no vale la pena leerlos todos -- igual va a terminar en
        # revisión manual por ambiguo. Cortar acá evita que un solo producto
        # trabe el proceso por varios minutos.
        if 0 < total <= MAX_CANDIDATOS_A_LEER:
            # Se extraen TODAS las filas en una sola llamada a JavaScript en
            # vez de recorrerlas una por una con Playwright -- con resultados
            # numerosos, hacer un viaje de ida y vuelta por cada celda es lo
            # que dejaba el proceso pegado varios minutos en un solo producto.
            filas_js = self.page.eval_on_selector_all(
                "#ctl00_ContentPlaceHolder1_gvDatosBusqueda tr",
                "trs => trs.slice(1).map(tr => "
                "Array.from(tr.querySelectorAll('td')).map(td => td.innerText.trim()))",
            )
            for celdas in filas_js:
                if len(celdas) >= 6:
                    filas.append(
                        {
                            "registro": celdas[1],
                            "nombre": celdas[2],
                            "fecha_registro": celdas[3],
                            "empresa": celdas[4],
                            "principio_activo": celdas[5],
                            "control_legal": celdas[6] if len(celdas) > 6 else "",
                        }
                    )
        return total, filas

    def buscar_por_nombre(self, nombre_producto):
        self._ir_a_busqueda_por_nombre()
        self.page.fill("#ctl00_ContentPlaceHolder1_txtNombreProducto", nombre_producto)
        self.page.click("#ctl00_ContentPlaceHolder1_btnBuscar")
        self.page.wait_for_load_state("networkidle", timeout=15000)
        return self._leer_resultados()

    def condicion_venta_de_registro(self, numero_registro):
        """Prueba los 5 valores oficiales hasta encontrar el que matchea este registro exacto."""
        self._ir_a_busqueda_por_registro_y_condicion()
        for condicion in CONDICIONES_VENTA:
            self.page.fill("#ctl00_ContentPlaceHolder1_txtNumeroRegistro", numero_registro)
            self.page.select_option(
                "#ctl00_ContentPlaceHolder1_ddlCondicion", label=condicion
            )
            self.page.click("#ctl00_ContentPlaceHolder1_btnBuscar")
            self.page.wait_for_load_state("networkidle", timeout=15000)
            total, _ = self._leer_resultados()
            if total >= 1:
                return condicion
            time.sleep(REQUEST_DELAY_SECONDS)
        return None


def con_reintentos(fn, *args, **kwargs):
    ultimo_error = None
    for intento in range(1, MAX_RETRIES + 1):
        try:
            return fn(*args, **kwargs)
        except (PWTimeout, Exception) as e:  # noqa: BLE001 -- script de uso único
            ultimo_error = e
            print(f"    reintento {intento}/{MAX_RETRIES} tras error: {e}")
            time.sleep(2 * intento)
    raise ultimo_error


def buscar_con_terminos_progresivos(cliente, nombre_producto):
    """Prueba primero con 1 palabra, y si no matchea nada (o es muy corta/
    genérica para buscar en serio, ej. "AB", "30"), con 2 -- ver
    limpiar_para_busqueda. Devuelve (termino_usado, total, candidatos)."""
    termino1 = limpiar_para_busqueda(nombre_producto, max_palabras=1)
    termino2 = limpiar_para_busqueda(nombre_producto, max_palabras=2)

    intentos = []
    if len(termino1) > 2 and not termino1.isdigit():
        intentos.append(termino1)
    if termino2 != termino1 and len(termino2) > 2:
        intentos.append(termino2)

    if not intentos:
        return termino1, 0, []

    for i, termino in enumerate(intentos):
        if i > 0:
            time.sleep(REQUEST_DELAY_SECONDS)
        total, candidatos = con_reintentos(cliente.buscar_por_nombre, termino)
        if total > 0:
            return termino, total, candidatos

    return intentos[-1], 0, []


def clasificar_producto(cliente, nombre_producto):
    """Devuelve un dict con las 4 columnas nuevas (ver COLUMNAS_NUEVAS)."""
    resultado = {
        "Es Medicamento": "NO",
        "Requiere Receta": "",
        "Revisar Manual": "NO",
        "Detalle": "",
    }

    termino, total, candidatos = buscar_con_terminos_progresivos(cliente, nombre_producto)

    if total == 0:
        if suena_a_medicamento(nombre_producto):
            resultado["Revisar Manual"] = "SI"
            resultado["Detalle"] = (
                f"No se encontró en el registro ISP buscando '{termino}', pero el "
                "nombre sugiere que podría ser un medicamento (revisar a mano; "
                "posible producto CENABAST, importado, o registrado con otro nombre)."
            )
        return resultado

    if total > MAX_CANDIDATOS_A_LEER:
        resultado["Es Medicamento"] = "SI (ambiguo)"
        resultado["Requiere Receta"] = "NO DETERMINADO"
        resultado["Revisar Manual"] = "SI"
        resultado["Detalle"] = (
            f"Búsqueda '{termino}' encontró {total} productos posibles -- demasiados "
            "para elegir uno solo automáticamente (término de búsqueda muy genérico). "
            "Revisar a mano en registrosanitario.ispch.gob.cl."
        )
        return resultado

    if total == 1:
        elegido = candidatos[0]
    else:
        dosis = extraer_dosis(nombre_producto)
        formas = extraer_formas(nombre_producto)
        filtrados = candidatos
        if dosis:
            # el ISP usa coma decimal ("27,5 mcg"), el inventario usa punto
            # ("27.5MCG") -- normalizar antes de comparar o el filtro no
            # matchea nunca.
            filtrados = [
                c for c in filtrados
                if dosis in c["nombre"].lower().replace(" ", "").replace(",", ".")
            ]
        if formas and len(filtrados) > 1:
            filtrados = [
                c for c in filtrados
                if any(f in c["nombre"].lower() for f in formas)
            ]
        if len(filtrados) == 1:
            elegido = filtrados[0]
        else:
            resultado["Es Medicamento"] = "SI (ambiguo)"
            resultado["Requiere Receta"] = "NO DETERMINADO"
            resultado["Revisar Manual"] = "SI"
            alternativas = " | ".join(
                f"{c['registro']}: {c['nombre']}" for c in candidatos[:5]
            )
            resultado["Detalle"] = (
                f"Búsqueda '{termino}' encontró {total} productos posibles, no se "
                f"pudo elegir uno solo: {alternativas}"
            )
            return resultado

    resultado["Es Medicamento"] = "SI"
    condicion = con_reintentos(cliente.condicion_venta_de_registro, elegido["registro"])
    if condicion:
        resultado["Requiere Receta"] = ETIQUETA_REQUIERE_RECETA[condicion]
    else:
        resultado["Requiere Receta"] = "NO DETERMINADO"
        resultado["Revisar Manual"] = "SI"

    resultado["Detalle"] = (
        f"Registro ISP {elegido['registro']} — {elegido['empresa']} — "
        f"Principio activo: {elegido['principio_activo'] or 'no informado'}"
    )
    return resultado


# Cada cuántos productos se reinicia el navegador de forma preventiva --
# evita que la memoria acumulada de miles de páginas termine tumbando el
# proceso a mitad de camino (lo que generaba filas de error en cadena).
RECICLAR_NAVEGADOR_CADA = 150
# Si el navegador se cae y se relanza más de esto SEGUIDO sin lograr
# procesar ni un producto, algo más grave está pasando (ej. un antivirus
# bloqueando chrome.exe) -- se corta en vez de reintentar para siempre.
MAX_RELANZAMIENTOS_SEGUIDOS = 5


def _crear_cliente(playwright):
    browser = playwright.chromium.launch()
    page = browser.new_page()
    return browser, RegistroSanitarioClient(page)


def main():
    if len(sys.argv) != 3:
        print("Uso: python clasificar_medicamentos.py <inventario.xlsx> <salida.csv>")
        sys.exit(1)

    path_excel, path_csv = sys.argv[1], sys.argv[2]
    header_original, filas = leer_inventario(path_excel)
    columna_codigo = "Código" if "Código" in header_original else header_original[0]

    ya_procesados = cargar_ya_procesados(path_csv, columna_codigo)
    pendientes = [f for f in filas if f["_codigo"] not in ya_procesados]

    print(f"Total inventario: {len(filas)}")
    print(f"Ya procesados (se saltan): {len(ya_procesados)}")
    print(f"Pendientes esta corrida: {len(pendientes)}")

    fieldnames = header_original + COLUMNAS_NUEVAS
    f_out, writer = abrir_csv_para_append(path_csv, fieldnames)

    with sync_playwright() as p:
        browser, cliente = _crear_cliente(p)
        items_desde_reinicio = 0
        relanzamientos_seguidos = 0

        try:
            i = 0
            while i < len(pendientes):
                fila_original = pendientes[i]
                codigo = fila_original["_codigo"]
                nombre = fila_original["_producto"]
                print(f"[{i + 1}/{len(pendientes)}] {codigo} — {nombre[:60]}")

                try:
                    nuevas_columnas = clasificar_producto(cliente, nombre)
                except Exception as e:  # noqa: BLE001
                    if parece_navegador_muerto(e):
                        relanzamientos_seguidos += 1
                        if relanzamientos_seguidos > MAX_RELANZAMIENTOS_SEGUIDOS:
                            print(
                                f"\nEl navegador se cerró {relanzamientos_seguidos} veces "
                                "seguidas sin lograr procesar ningún producto -- algo más "
                                "grave está pasando (revisar antivirus/Defender). Se corta "
                                "acá; ya quedó guardado el avance, se puede retomar más tarde "
                                "con el mismo comando."
                            )
                            break
                        print(
                            f"    El navegador se cerró/crasheó ({e}). Reiniciando "
                            "navegador y reintentando este mismo producto (no se pierde)…"
                        )
                        try:
                            browser.close()
                        except Exception:  # noqa: BLE001
                            pass
                        browser, cliente = _crear_cliente(p)
                        items_desde_reinicio = 0
                        continue  # reintenta el MISMO producto, no avanza i ni escribe fila

                    print(f"    FALLÓ tras reintentos, se marca para revisión manual: {e}")
                    nuevas_columnas = {
                        "Es Medicamento": "", "Requiere Receta": "",
                        "Revisar Manual": "SI",
                        "Detalle": f"Error al consultar el ISP: {e}",
                    }

                relanzamientos_seguidos = 0
                fila_salida = {k: fila_original[k] for k in header_original}
                fila_salida.update(nuevas_columnas)
                writer.writerow(fila_salida)
                f_out.flush()
                i += 1
                items_desde_reinicio += 1

                if items_desde_reinicio >= RECICLAR_NAVEGADOR_CADA:
                    print("    (reiniciando navegador de forma preventiva por memoria)")
                    browser.close()
                    browser, cliente = _crear_cliente(p)
                    items_desde_reinicio = 0

                time.sleep(REQUEST_DELAY_SECONDS)
        except KeyboardInterrupt:
            print("\nInterrumpido por el usuario. Progreso guardado en", path_csv)
        finally:
            try:
                browser.close()
            except Exception:  # noqa: BLE001
                pass
            f_out.close()


if __name__ == "__main__":
    main()
